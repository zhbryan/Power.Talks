"""
ERCOT RMGRR Document Downloader
================================
Automatically fetches all *pending*, *withdrawn*, and *approved* Retail
Market Guide Revision Requests (RMGRRs) from the ERCOT website, checks which
documents are already saved locally, and downloads only the new ones
(PDF, DOC, DOCX, XLS, XLSX).

Usage
-----
    pip install requests beautifulsoup4 openpyxl
    python download_ercot_rmgrr.py

Configuration
-------------
Edit the SETTINGS block below to change where files are saved, or to
override the auto-fetched lists with specific RMGRR numbers.
"""

import os
import re
import time
import requests
import openpyxl
from bs4 import BeautifulSoup
from datetime import datetime
from urllib.parse import urljoin, unquote

# ── SETTINGS ─────────────────────────────────────────────────────────────────

# Override: set to a list of RMGRR numbers to process instead of auto-fetching.
# Applies to pending, withdrawn, and approved runs.  Leave as None for auto-fetch.
# Example: RMGRR_NUMBERS = [100, 110, 120]
RMGRR_NUMBERS = None

# Only process approved RMGRRs with a number greater than this value.
APPROVED_MIN_RMGRR = 0

# Local folder where files will be saved (one sub-folder per RMGRR).
PENDING_DIR   = r"E:\wamp64\www\Power.Talks\Documents Database\ERCOT.MKT.RULES\RMGRR"
WITHDRAWN_DIR = PENDING_DIR
APPROVED_DIR  = PENDING_DIR

# Seconds to pause between HTTP requests (be polite to the server).
REQUEST_DELAY = 1.5

# File extensions to download.
DOWNLOAD_EXTS = {".pdf", ".doc", ".docx", ".xls", ".xlsx"}

# Excel workbook used to track which RMGRR numbers have been seen before.
EXCEL_TRACKER = r"E:\wamp64\www\Power.Talks\Documents Database\ERCOT.MKT.RULES\RMGRR\Current List of RMGRRs.xlsx"

# ─────────────────────────────────────────────────────────────────────────────

BASE_URL            = "https://www.ercot.com"
RMGRR_LIST_URL      = "https://www.ercot.com/mktrules/issues/rmgrr"
RMGRR_PENDING_URL   = "https://www.ercot.com/mktrules/issues/reports/rmgrr/pending"
RMGRR_WITHDRAWN_URL = "https://www.ercot.com/mktrules/issues/reports/rmgrr/withdrawn"
RMGRR_APPROVED_URL  = "https://www.ercot.com/mktrules/issues/reports/rmgrr/approved"
ISSUE_URL           = "https://www.ercot.com/mktrules/issues/RMGRR{n}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; RMGRR-Downloader/1.0; "
        "non-commercial research)"
    )
}

MAX_PATH = 240  # conservative Windows path-length limit


def sanitize(name: str) -> str:
    """Strip characters that are illegal in file/folder names."""
    return re.sub(r'[\\/*?:"<>|]', "_", name).strip()


def safe_fname(fname: str, dest_dir: str) -> str:
    """Truncate fname so the full path stays within MAX_PATH."""
    stem, ext = os.path.splitext(fname)
    max_stem = MAX_PATH - len(dest_dir) - len(os.sep) - len(ext)
    if max_stem < 1:
        max_stem = 1
    return stem[:max_stem] + ext


def fetch_rmgrr_numbers_from_report(url: str, label: str) -> list[int]:
    """Scrape a report page and return a sorted list of RMGRR numbers."""
    print(f"Fetching {label} RMGRR list from {url} ...")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [ERR] Could not fetch {label} RMGRR list: {exc}")
        return []

    numbers = re.findall(r"/mktrules/issues/RMGRR(\d+)", resp.text)
    return sorted(set(int(n) for n in numbers))


def get_pending_rmgrrs() -> list[int]:
    """
    Scrape the ERCOT RMGRR listing page for the Pending section.
    Falls back to the pending report page if the section is not found.
    """
    print(f"Fetching pending RMGRR list from {RMGRR_LIST_URL} ...")
    try:
        resp = requests.get(RMGRR_LIST_URL, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [ERR] Could not fetch RMGRR list: {exc}")
        return fetch_rmgrr_numbers_from_report(RMGRR_PENDING_URL, "pending")

    soup = BeautifulSoup(resp.text, "html.parser")

    pending_heading = soup.find(
        lambda tag: tag.name in ("h3", "h4", "h5")
        and "pending" in tag.get_text(strip=True).lower()
    )
    if not pending_heading:
        print("  [WARN] Could not find 'Pending' section on listing page — falling back to report page.")
        return fetch_rmgrr_numbers_from_report(RMGRR_PENDING_URL, "pending")

    rmgrr_numbers = []
    for sibling in pending_heading.find_next_siblings():
        if sibling.name in ("h3", "h4", "h5"):
            break
        for a in sibling.find_all("a", href=True):
            m = re.search(r"/mktrules/issues/RMGRR(\d+)", a["href"])
            if m:
                rmgrr_numbers.append(int(m.group(1)))

    if not rmgrr_numbers:
        print("  [WARN] No RMGRRs found in Pending section — falling back to report page.")
        return fetch_rmgrr_numbers_from_report(RMGRR_PENDING_URL, "pending")

    seen, unique = set(), []
    for n in rmgrr_numbers:
        if n not in seen:
            seen.add(n)
            unique.append(n)
    return unique


def get_document_links(rmgrr: int) -> list[dict]:
    """
    Fetch the ERCOT issue page for one RMGRR and return a list of dicts
    {'label': str, 'url': str} for every downloadable file found.
    Returns an empty list if the page is unreachable or has no files.
    """
    url = ISSUE_URL.format(n=rmgrr)
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        if resp.status_code == 404:
            return []
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [WARN] Could not fetch {url}: {exc}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    links, seen = [], set()

    for a in soup.find_all("a", href=True):
        href = a["href"]
        ext  = os.path.splitext(href.split("?")[0])[-1].lower()
        if ext not in DOWNLOAD_EXTS:
            continue
        full_url = urljoin(BASE_URL, href)
        if full_url in seen:
            continue
        seen.add(full_url)
        label = a.get_text(separator=" ", strip=True) or os.path.basename(href)
        links.append({"label": label, "url": full_url})

    return links


def download_file(url: str, dest_path: str) -> bool:
    """Download *url* to *dest_path*. Returns True on success.

    Writes to a temporary file first and renames on success so that a failed
    or interrupted download never leaves a corrupt file that would be mistaken
    for a completed download on the next run.
    """
    tmp_path = dest_path + ".tmp"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=60, stream=True)
        resp.raise_for_status()
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        with open(tmp_path, "wb") as fh:
            for chunk in resp.iter_content(chunk_size=65536):
                fh.write(chunk)
        os.replace(tmp_path, dest_path)
        size_kb = os.path.getsize(dest_path) / 1024
        print(f"    [OK]   {os.path.basename(dest_path)}  ({size_kb:.1f} KB)")
        return True
    except (requests.RequestException, OSError) as exc:
        print(f"    [ERR]  Failed to download {url}: {exc}")
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        return False


def process_rmgrr_list(rmgrr_list: list[int], output_dir: str) -> tuple[int, int]:
    """
    For each RMGRR in *rmgrr_list*, check for new documents and download them
    into *output_dir*/RMGRR{n}/.  Returns (files_downloaded, rmgrrs_updated).
    """
    total_files = total_rmgrrs = 0

    for rmgrr in rmgrr_list:
        print(f"\nRMGRR{rmgrr}  ->  {ISSUE_URL.format(n=rmgrr)}")
        links = get_document_links(rmgrr)
        time.sleep(REQUEST_DELAY)

        if not links:
            print("  (no downloadable documents found or page does not exist)")
            continue

        rmgrr_dir = os.path.join(output_dir, f"RMGRR{rmgrr}")

        new_items = []
        for item in links:
            fname = sanitize(unquote(os.path.basename(item["url"].split("?")[0])))
            if not fname:
                fname = sanitize(item["label"]) + ".bin"
            fname = safe_fname(fname, rmgrr_dir)
            dest  = os.path.join(rmgrr_dir, fname)
            if not os.path.exists(dest):
                new_items.append((item, fname, dest))

        already = len(links) - len(new_items)
        if already:
            print(f"  {already} file(s) already up to date.")

        if not new_items:
            print("  Nothing new to download.")
            continue

        print(f"  {len(new_items)} new file(s) to download "
              f"(out of {len(links)} total).")
        os.makedirs(rmgrr_dir, exist_ok=True)
        total_rmgrrs += 1

        for item, fname, dest in new_items:
            if download_file(item["url"], dest):
                total_files += 1
            time.sleep(REQUEST_DELAY)

    return total_files, total_rmgrrs


def load_excel_rmgrrs(sheet_name: str) -> set[int]:
    """
    Read RMGRR numbers from *sheet_name* in EXCEL_TRACKER.
    Returns an empty set if the file or sheet does not exist.
    """
    if not os.path.exists(EXCEL_TRACKER):
        return set()
    try:
        wb = openpyxl.load_workbook(EXCEL_TRACKER, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return set()
        ws = wb[sheet_name]
        numbers = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                try:
                    numbers.add(int(row[0]))
                except (ValueError, TypeError):
                    pass
        wb.close()
        return numbers
    except Exception as exc:
        print(f"  [WARN] Could not read {EXCEL_TRACKER}: {exc}")
        return set()


def update_excel_rmgrrs(sheet_name: str, rmgrr_list: list[int]) -> None:
    """
    Overwrite *sheet_name* in EXCEL_TRACKER with the full *rmgrr_list*,
    sorted ascending.  Creates the workbook/sheet if they do not exist.
    """
    os.makedirs(os.path.dirname(EXCEL_TRACKER), exist_ok=True)
    if os.path.exists(EXCEL_TRACKER):
        wb = openpyxl.load_workbook(EXCEL_TRACKER)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append(["RMGRR_Number", f"Last Updated: {timestamp}"])
    for n in sorted(rmgrr_list):
        ws.append([n])

    wb.save(EXCEL_TRACKER)
    print(f"  [Excel] '{sheet_name}' updated — {len(rmgrr_list)} RMGRR(s) saved to {EXCEL_TRACKER} (as of {timestamp})")


def main():
    # ── Pending ───────────────────────────────────────────────────────────────
    if RMGRR_NUMBERS:
        pending_list = RMGRR_NUMBERS
        print("ERCOT RMGRR Document Downloader  (manual list)")
    else:
        pending_list = get_pending_rmgrrs()
        if not pending_list:
            print("No pending RMGRRs found.")
        else:
            print(f"Found {len(pending_list)} pending RMGRR(s): "
                  f"{pending_list[0]} - {pending_list[-1]}")

    # ── Withdrawn ─────────────────────────────────────────────────────────────
    withdrawn_list = RMGRR_NUMBERS if RMGRR_NUMBERS else fetch_rmgrr_numbers_from_report(RMGRR_WITHDRAWN_URL, "withdrawn")
    withdrawn_new = withdrawn_list
    if withdrawn_list:
        print(f"Found {len(withdrawn_list)} withdrawn RMGRR(s): "
              f"{withdrawn_list[0]} - {withdrawn_list[-1]}")
        tracked_withdrawn = load_excel_rmgrrs("List_Withdrawn")
        if tracked_withdrawn:
            withdrawn_new = [n for n in withdrawn_list if n not in tracked_withdrawn]
            print(f"  {len(tracked_withdrawn)} already tracked in Excel, "
                  f"{len(withdrawn_new)} new RMGRR(s) to process.")
        else:
            print("  No prior tracking found — will process all withdrawn RMGRRs.")

    # ── Approved ──────────────────────────────────────────────────────────────
    approved_list = RMGRR_NUMBERS if RMGRR_NUMBERS else fetch_rmgrr_numbers_from_report(RMGRR_APPROVED_URL, "approved")
    approved_list = [n for n in approved_list if n > APPROVED_MIN_RMGRR]
    approved_new = approved_list
    if approved_list:
        print(f"Found {len(approved_list)} approved RMGRR(s) > {APPROVED_MIN_RMGRR}: "
              f"{approved_list[0]} - {approved_list[-1]}")
        tracked_approved = load_excel_rmgrrs("List_Approved")
        if tracked_approved:
            approved_new = [n for n in approved_list if n not in tracked_approved]
            print(f"  {len(tracked_approved)} already tracked in Excel, "
                  f"{len(approved_new)} new RMGRR(s) to process.")
        else:
            print("  No prior tracking found — will process all approved RMGRRs.")

    print("=" * 60)

    # ── Process pending ───────────────────────────────────────────────────────
    p_files = p_rmgrrs = 0
    if pending_list:
        print(f"\n{'='*60}")
        print(f"PENDING RMGRRs  ->  {os.path.abspath(PENDING_DIR)}")
        print(f"{'='*60}")
        p_files, p_rmgrrs = process_rmgrr_list(pending_list, PENDING_DIR)
        update_excel_rmgrrs("List_Pending", pending_list)

    # ── Process withdrawn ─────────────────────────────────────────────────────
    w_files = w_rmgrrs = 0
    if withdrawn_list:
        print(f"\n{'='*60}")
        print(f"WITHDRAWN RMGRRs  ->  {os.path.abspath(WITHDRAWN_DIR)}")
        print(f"{'='*60}")
        w_files, w_rmgrrs = process_rmgrr_list(withdrawn_new, WITHDRAWN_DIR)
        update_excel_rmgrrs("List_Withdrawn", withdrawn_list)

    # ── Process approved ──────────────────────────────────────────────────────
    a_files = a_rmgrrs = 0
    if approved_list:
        print(f"\n{'='*60}")
        print(f"APPROVED RMGRRs  ->  {os.path.abspath(APPROVED_DIR)}")
        print(f"{'='*60}")
        a_files, a_rmgrrs = process_rmgrr_list(approved_new, APPROVED_DIR)
        update_excel_rmgrrs("List_Approved", approved_list)

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print(f"Pending  : {p_files} new file(s) across {p_rmgrrs} RMGRR(s)  ->  {os.path.abspath(PENDING_DIR)}")
    print(f"Withdrawn: {w_files} new file(s) across {w_rmgrrs} RMGRR(s)  ->  {os.path.abspath(WITHDRAWN_DIR)}")
    print(f"Approved : {a_files} new file(s) across {a_rmgrrs} RMGRR(s)  ->  {os.path.abspath(APPROVED_DIR)}")
    print(f"Total    : {p_files + w_files + a_files} new file(s) downloaded.")


if __name__ == "__main__":
    main()

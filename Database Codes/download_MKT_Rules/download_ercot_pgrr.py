"""
ERCOT PGRR Document Downloader
================================
Automatically fetches all *pending*, *withdrawn*, and *approved* Planning
Guide Revision Requests (PGRRs) from the ERCOT website, checks which
documents are already saved locally, and downloads only the new ones
(PDF, DOC, DOCX, XLS, XLSX).

Usage
-----
    pip install requests beautifulsoup4 openpyxl
    python download_ercot_pgrr.py

Configuration
-------------
Edit the SETTINGS block below to change where files are saved, or to
override the auto-fetched lists with specific PGRR numbers.
"""

import os
import re
import time
import requests
import openpyxl
from bs4 import BeautifulSoup
from datetime import datetime
from urllib.parse import urljoin, unquote

# ── SETTINGS ─────────────────────────────────────────────────────────────────

# Override: set to a list of PGRR numbers to process instead of auto-fetching.
# Applies to pending, withdrawn, and approved runs.  Leave as None for auto-fetch.
# Example: PGRR_NUMBERS = [100, 110, 120]
PGRR_NUMBERS = None

# Only process approved PGRRs with a number greater than this value.
APPROVED_MIN_PGRR = 0

# Local folder where files will be saved (one sub-folder per PGRR).
PENDING_DIR   = r"E:\wamp64\www\Power.Talks\Documents Database\ERCOT.MKT.RULES\PGRR"
WITHDRAWN_DIR = PENDING_DIR
APPROVED_DIR  = PENDING_DIR

# Seconds to pause between HTTP requests (be polite to the server).
REQUEST_DELAY = 1.5

# File extensions to download.
DOWNLOAD_EXTS = {".pdf", ".doc", ".docx", ".xls", ".xlsx"}

# Excel workbook used to track which PGRR numbers have been seen before.
EXCEL_TRACKER = r"E:\wamp64\www\Power.Talks\Documents Database\ERCOT.MKT.RULES\PGRR\Current List of PGRRs.xlsx"

# ─────────────────────────────────────────────────────────────────────────────

BASE_URL           = "https://www.ercot.com"
PGRR_LIST_URL      = "https://www.ercot.com/mktrules/issues/pgrr"
PGRR_WITHDRAWN_URL = "https://www.ercot.com/mktrules/issues/reports/pgrr/withdrawn"
PGRR_APPROVED_URL  = "https://www.ercot.com/mktrules/issues/reports/pgrr/approved"
ISSUE_URL          = "https://www.ercot.com/mktrules/issues/PGRR{n}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; PGRR-Downloader/1.0; "
        "non-commercial research)"
    )
}

MAX_PATH = 240  # conservative Windows path-length limit


def sanitize(name: str) -> str:
    """Strip characters that are illegal in file/folder names."""
    return re.sub(r'[\\/*?:"<>|]', "_", name).strip()


def safe_fname(fname: str, dest_dir: str) -> str:
    """Truncate fname so the full path stays within MAX_PATH."""
    stem, ext = os.path.splitext(fname)
    max_stem = MAX_PATH - len(dest_dir) - len(os.sep) - len(ext)
    if max_stem < 1:
        max_stem = 1
    return stem[:max_stem] + ext


def get_pending_pgrrs() -> list[int]:
    """
    Scrape the ERCOT PGRR listing page and return PGRR numbers listed
    under the 'Pending' section.
    """
    print(f"Fetching pending PGRR list from {PGRR_LIST_URL} ...")
    try:
        resp = requests.get(PGRR_LIST_URL, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [ERR] Could not fetch PGRR list: {exc}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")

    pending_heading = soup.find(
        lambda tag: tag.name in ("h3", "h4", "h5")
        and "pending" in tag.get_text(strip=True).lower()
    )
    if not pending_heading:
        print("  [WARN] Could not find 'Pending' section on listing page.")
        return []

    pgrr_numbers = []
    for sibling in pending_heading.find_next_siblings():
        if sibling.name in ("h3", "h4", "h5"):
            break
        for a in sibling.find_all("a", href=True):
            m = re.search(r"/mktrules/issues/PGRR(\d+)", a["href"])
            if m:
                pgrr_numbers.append(int(m.group(1)))

    seen, unique = set(), []
    for n in pgrr_numbers:
        if n not in seen:
            seen.add(n)
            unique.append(n)
    return unique


def get_withdrawn_pgrrs() -> list[int]:
    """
    Scrape the ERCOT withdrawn-PGRR report page and return a sorted list
    of withdrawn PGRR numbers.
    """
    print(f"Fetching withdrawn PGRR list from {PGRR_WITHDRAWN_URL} ...")
    try:
        resp = requests.get(PGRR_WITHDRAWN_URL, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [ERR] Could not fetch withdrawn PGRR list: {exc}")
        return []

    numbers = re.findall(r"/mktrules/issues/PGRR(\d+)", resp.text)
    return sorted(set(int(n) for n in numbers))


def get_approved_pgrrs() -> list[int]:
    """
    Scrape the ERCOT approved-PGRR report page and return a sorted list
    of approved PGRR numbers.
    """
    print(f"Fetching approved PGRR list from {PGRR_APPROVED_URL} ...")
    try:
        resp = requests.get(PGRR_APPROVED_URL, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [ERR] Could not fetch approved PGRR list: {exc}")
        return []

    numbers = re.findall(r"/mktrules/issues/PGRR(\d+)", resp.text)
    return sorted(set(int(n) for n in numbers))


def get_document_links(pgrr: int) -> list[dict]:
    """
    Fetch the ERCOT issue page for one PGRR and return a list of dicts
    {'label': str, 'url': str} for every downloadable file found.
    Returns an empty list if the page is unreachable or has no files.
    """
    url = ISSUE_URL.format(n=pgrr)
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        if resp.status_code == 404:
            return []
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [WARN] Could not fetch {url}: {exc}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    links, seen = [], set()

    for a in soup.find_all("a", href=True):
        href = a["href"]
        ext  = os.path.splitext(href.split("?")[0])[-1].lower()
        if ext not in DOWNLOAD_EXTS:
            continue
        full_url = urljoin(BASE_URL, href)
        if full_url in seen:
            continue
        seen.add(full_url)
        label = a.get_text(separator=" ", strip=True) or os.path.basename(href)
        links.append({"label": label, "url": full_url})

    return links


def download_file(url: str, dest_path: str) -> bool:
    """Download *url* to *dest_path*. Returns True on success.

    Writes to a temporary file first and renames on success so that a failed
    or interrupted download never leaves a corrupt file that would be mistaken
    for a completed download on the next run.
    """
    tmp_path = dest_path + ".tmp"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=60, stream=True)
        resp.raise_for_status()
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        with open(tmp_path, "wb") as fh:
            for chunk in resp.iter_content(chunk_size=65536):
                fh.write(chunk)
        os.replace(tmp_path, dest_path)
        size_kb = os.path.getsize(dest_path) / 1024
        print(f"    [OK]   {os.path.basename(dest_path)}  ({size_kb:.1f} KB)")
        return True
    except (requests.RequestException, OSError) as exc:
        print(f"    [ERR]  Failed to download {url}: {exc}")
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        return False


def process_pgrr_list(pgrr_list: list[int], output_dir: str) -> tuple[int, int]:
    """
    For each PGRR in *pgrr_list*, check for new documents and download them
    into *output_dir*/PGRR{n}/.  Returns (files_downloaded, pgrrs_updated).
    """
    total_files = total_pgrrs = 0

    for pgrr in pgrr_list:
        print(f"\nPGRR{pgrr}  ->  {ISSUE_URL.format(n=pgrr)}")
        links = get_document_links(pgrr)
        time.sleep(REQUEST_DELAY)

        if not links:
            print("  (no downloadable documents found or page does not exist)")
            continue

        pgrr_dir = os.path.join(output_dir, f"PGRR{pgrr}")

        new_items = []
        for item in links:
            fname = sanitize(unquote(os.path.basename(item["url"].split("?")[0])))
            if not fname:
                fname = sanitize(item["label"]) + ".bin"
            fname = safe_fname(fname, pgrr_dir)
            dest  = os.path.join(pgrr_dir, fname)
            if not os.path.exists(dest):
                new_items.append((item, fname, dest))

        already = len(links) - len(new_items)
        if already:
            print(f"  {already} file(s) already up to date.")

        if not new_items:
            print("  Nothing new to download.")
            continue

        print(f"  {len(new_items)} new file(s) to download "
              f"(out of {len(links)} total).")
        os.makedirs(pgrr_dir, exist_ok=True)
        total_pgrrs += 1

        for item, fname, dest in new_items:
            if download_file(item["url"], dest):
                total_files += 1
            time.sleep(REQUEST_DELAY)

    return total_files, total_pgrrs


def load_excel_pgrrs(sheet_name: str) -> set[int]:
    """
    Read PGRR numbers from *sheet_name* in EXCEL_TRACKER.
    Returns an empty set if the file or sheet does not exist.
    """
    if not os.path.exists(EXCEL_TRACKER):
        return set()
    try:
        wb = openpyxl.load_workbook(EXCEL_TRACKER, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return set()
        ws = wb[sheet_name]
        numbers = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                try:
                    numbers.add(int(row[0]))
                except (ValueError, TypeError):
                    pass
        wb.close()
        return numbers
    except Exception as exc:
        print(f"  [WARN] Could not read {EXCEL_TRACKER}: {exc}")
        return set()


def update_excel_pgrrs(sheet_name: str, pgrr_list: list[int]) -> None:
    """
    Overwrite *sheet_name* in EXCEL_TRACKER with the full *pgrr_list*,
    sorted ascending.  Creates the workbook/sheet if they do not exist.
    """
    os.makedirs(os.path.dirname(EXCEL_TRACKER), exist_ok=True)
    if os.path.exists(EXCEL_TRACKER):
        wb = openpyxl.load_workbook(EXCEL_TRACKER)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append(["PGRR_Number", f"Last Updated: {timestamp}"])
    for n in sorted(pgrr_list):
        ws.append([n])

    wb.save(EXCEL_TRACKER)
    print(f"  [Excel] '{sheet_name}' updated — {len(pgrr_list)} PGRR(s) saved to {EXCEL_TRACKER} (as of {timestamp})")


def main():
    # ── Pending ───────────────────────────────────────────────────────────────
    if PGRR_NUMBERS:
        pending_list = PGRR_NUMBERS
        print("ERCOT PGRR Document Downloader  (manual list)")
    else:
        pending_list = get_pending_pgrrs()
        if not pending_list:
            print("No pending PGRRs found.")
        else:
            print(f"Found {len(pending_list)} pending PGRR(s): "
                  f"{pending_list[0]} - {pending_list[-1]}")

    # ── Withdrawn ─────────────────────────────────────────────────────────────
    withdrawn_list = PGRR_NUMBERS if PGRR_NUMBERS else get_withdrawn_pgrrs()
    withdrawn_new = withdrawn_list
    if withdrawn_list:
        print(f"Found {len(withdrawn_list)} withdrawn PGRR(s): "
              f"{withdrawn_list[0]} - {withdrawn_list[-1]}")
        tracked_withdrawn = load_excel_pgrrs("List_Withdrawn")
        if tracked_withdrawn:
            withdrawn_new = [n for n in withdrawn_list if n not in tracked_withdrawn]
            print(f"  {len(tracked_withdrawn)} already tracked in Excel, "
                  f"{len(withdrawn_new)} new PGRR(s) to process.")
        else:
            print("  No prior tracking found — will process all withdrawn PGRRs.")

    # ── Approved ──────────────────────────────────────────────────────────────
    approved_list = PGRR_NUMBERS if PGRR_NUMBERS else get_approved_pgrrs()
    approved_list = [n for n in approved_list if n > APPROVED_MIN_PGRR]
    approved_new = approved_list
    if approved_list:
        print(f"Found {len(approved_list)} approved PGRR(s) > {APPROVED_MIN_PGRR}: "
              f"{approved_list[0]} - {approved_list[-1]}")
        tracked_approved = load_excel_pgrrs("List_Approved")
        if tracked_approved:
            approved_new = [n for n in approved_list if n not in tracked_approved]
            print(f"  {len(tracked_approved)} already tracked in Excel, "
                  f"{len(approved_new)} new PGRR(s) to process.")
        else:
            print("  No prior tracking found — will process all approved PGRRs.")

    print("=" * 60)

    # ── Process pending ───────────────────────────────────────────────────────
    p_files = p_pgrrs = 0
    if pending_list:
        print(f"\n{'='*60}")
        print(f"PENDING PGRRs  ->  {os.path.abspath(PENDING_DIR)}")
        print(f"{'='*60}")
        p_files, p_pgrrs = process_pgrr_list(pending_list, PENDING_DIR)
        update_excel_pgrrs("List_Pending", pending_list)

    # ── Process withdrawn ─────────────────────────────────────────────────────
    w_files = w_pgrrs = 0
    if withdrawn_list:
        print(f"\n{'='*60}")
        print(f"WITHDRAWN PGRRs  ->  {os.path.abspath(WITHDRAWN_DIR)}")
        print(f"{'='*60}")
        w_files, w_pgrrs = process_pgrr_list(withdrawn_new, WITHDRAWN_DIR)
        update_excel_pgrrs("List_Withdrawn", withdrawn_list)

    # ── Process approved ──────────────────────────────────────────────────────
    a_files = a_pgrrs = 0
    if approved_list:
        print(f"\n{'='*60}")
        print(f"APPROVED PGRRs  ->  {os.path.abspath(APPROVED_DIR)}")
        print(f"{'='*60}")
        a_files, a_pgrrs = process_pgrr_list(approved_new, APPROVED_DIR)
        update_excel_pgrrs("List_Approved", approved_list)

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print(f"Pending  : {p_files} new file(s) across {p_pgrrs} PGRR(s)  ->  {os.path.abspath(PENDING_DIR)}")
    print(f"Withdrawn: {w_files} new file(s) across {w_pgrrs} PGRR(s)  ->  {os.path.abspath(WITHDRAWN_DIR)}")
    print(f"Approved : {a_files} new file(s) across {a_pgrrs} PGRR(s)  ->  {os.path.abspath(APPROVED_DIR)}")
    print(f"Total    : {p_files + w_files + a_files} new file(s) downloaded.")


if __name__ == "__main__":
    main()

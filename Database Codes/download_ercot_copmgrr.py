"""
ERCOT COPMGRR Document Downloader
===================================
Automatically fetches all *pending*, *withdrawn*, and *approved* COPMGRRs
from the ERCOT website, checks which documents are already saved locally,
and downloads only the new ones (PDF, DOC, DOCX, XLS, XLSX).

Usage
-----
    pip install requests beautifulsoup4 openpyxl
    python download_ercot_copmgrr.py

Configuration
-------------
Edit the SETTINGS block below to change where files are saved, or to
override the auto-fetched lists with specific COPMGRR numbers.
"""

import os
import re
import time
import requests
import openpyxl
from bs4 import BeautifulSoup
from datetime import datetime
from urllib.parse import urljoin, unquote

# ── SETTINGS ─────────────────────────────────────────────────────────────────

# Override: set to a list of COPMGRR numbers to process instead of auto-fetching.
# Applies to pending, withdrawn, and approved runs.  Leave as None for auto-fetch.
# Example: COPMGRR_NUMBERS = [10, 15, 20]
COPMGRR_NUMBERS = None

# Only process approved COPMGRRs with a number greater than this value.
# Set higher if early approved numbers have no downloadable documents on ERCOT's site.
APPROVED_MIN_COPMGRR = 0


# Local folder where files will be saved (one sub-folder per COPMGRR).
PENDING_DIR   = r"E:\wamp64\www\Power.Talks\Documents Database\ERCOT.MKT.RULES\COPMGRR"
WITHDRAWN_DIR = PENDING_DIR
APPROVED_DIR  = PENDING_DIR

# Seconds to pause between HTTP requests (be polite to the server).
REQUEST_DELAY = 1.5

# File extensions to download.
DOWNLOAD_EXTS = {".pdf", ".doc", ".docx", ".xls", ".xlsx"}

# Excel workbook used to track which COPMGRR numbers have been seen before.
EXCEL_TRACKER = r"E:\wamp64\www\Power.Talks\Documents Database\ERCOT.MKT.RULES\COPMGRR\Current List of COPMGRRs.xlsx"

# ─────────────────────────────────────────────────────────────────────────────

BASE_URL              = "https://www.ercot.com"
COPMGRR_LIST_URL      = "https://www.ercot.com/mktrules/issues/copmgrr"
COPMGRR_WITHDRAWN_URL = "https://www.ercot.com/mktrules/issues/reports/copmgrr/withdrawn"
COPMGRR_APPROVED_URL  = "https://www.ercot.com/mktrules/issues/reports/copmgrr/approved"
ISSUE_URL             = "https://www.ercot.com/mktrules/issues/COPMGRR{n}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; COPMGRR-Downloader/1.0; "
        "non-commercial research)"
    )
}

MAX_PATH = 240  # conservative Windows path-length limit


def sanitize(name: str) -> str:
    """Strip characters that are illegal in file/folder names."""
    return re.sub(r'[\\/*?:"<>|]', "_", name).strip()


def safe_fname(fname: str, dest_dir: str) -> str:
    """Truncate fname so the full path stays within MAX_PATH."""
    stem, ext = os.path.splitext(fname)
    max_stem = MAX_PATH - len(dest_dir) - len(os.sep) - len(ext)
    if max_stem < 1:
        max_stem = 1
    return stem[:max_stem] + ext


def get_pending_copmgrrs() -> list[int]:
    """
    Scrape the ERCOT COPMGRR listing page and return COPMGRR numbers listed
    under the 'Pending' section.
    """
    print(f"Fetching pending COPMGRR list from {COPMGRR_LIST_URL} ...")
    try:
        resp = requests.get(COPMGRR_LIST_URL, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [ERR] Could not fetch COPMGRR list: {exc}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")

    pending_heading = soup.find(
        lambda tag: tag.name in ("h3", "h4", "h5")
        and "pending" in tag.get_text(strip=True).lower()
    )
    if not pending_heading:
        print("  [WARN] Could not find 'Pending' section on listing page.")
        return []

    copmgrr_numbers = []
    for sibling in pending_heading.find_next_siblings():
        if sibling.name in ("h3", "h4", "h5"):
            break
        for a in sibling.find_all("a", href=True):
            m = re.search(r"/mktrules/issues/COPMGRR(\d+)", a["href"])
            if m:
                copmgrr_numbers.append(int(m.group(1)))

    seen, unique = set(), []
    for n in copmgrr_numbers:
        if n not in seen:
            seen.add(n)
            unique.append(n)
    return unique


def get_withdrawn_copmgrrs() -> list[int]:
    """
    Scrape the ERCOT withdrawn-COPMGRR report page and return a sorted list
    of withdrawn COPMGRR numbers.
    """
    print(f"Fetching withdrawn COPMGRR list from {COPMGRR_WITHDRAWN_URL} ...")
    try:
        resp = requests.get(COPMGRR_WITHDRAWN_URL, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [ERR] Could not fetch withdrawn COPMGRR list: {exc}")
        return []

    numbers = re.findall(r"/mktrules/issues/COPMGRR(\d+)", resp.text)
    return sorted(set(int(n) for n in numbers))


def get_approved_copmgrrs() -> list[int]:
    """
    Scrape the ERCOT approved-COPMGRR report page and return a sorted list
    of approved COPMGRR numbers.
    """
    print(f"Fetching approved COPMGRR list from {COPMGRR_APPROVED_URL} ...")
    try:
        resp = requests.get(COPMGRR_APPROVED_URL, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [ERR] Could not fetch approved COPMGRR list: {exc}")
        return []

    numbers = re.findall(r"/mktrules/issues/COPMGRR(\d+)", resp.text)
    return sorted(set(int(n) for n in numbers))


def get_document_links(copmgrr: int) -> list[dict]:
    """
    Fetch the ERCOT issue page for one COPMGRR and return a list of dicts
    {'label': str, 'url': str} for every downloadable file found.
    Returns an empty list if the page is unreachable or has no files.
    """
    url = ISSUE_URL.format(n=copmgrr)
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        if resp.status_code == 404:
            return []
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [WARN] Could not fetch {url}: {exc}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    links, seen = [], set()

    for a in soup.find_all("a", href=True):
        href = a["href"]
        ext  = os.path.splitext(href.split("?")[0])[-1].lower()
        if ext not in DOWNLOAD_EXTS:
            continue
        full_url = urljoin(BASE_URL, href)
        if full_url in seen:
            continue
        seen.add(full_url)
        label = a.get_text(separator=" ", strip=True) or os.path.basename(href)
        links.append({"label": label, "url": full_url})

    return links


def download_file(url: str, dest_path: str) -> bool:
    """Download *url* to *dest_path*. Returns True on success.

    Writes to a temporary file first and renames on success so that a failed
    or interrupted download never leaves a corrupt file that would be mistaken
    for a completed download on the next run.
    """
    tmp_path = dest_path + ".tmp"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=60, stream=True)
        resp.raise_for_status()
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        with open(tmp_path, "wb") as fh:
            for chunk in resp.iter_content(chunk_size=65536):
                fh.write(chunk)
        os.replace(tmp_path, dest_path)
        size_kb = os.path.getsize(dest_path) / 1024
        print(f"    [OK]   {os.path.basename(dest_path)}  ({size_kb:.1f} KB)")
        return True
    except (requests.RequestException, OSError) as exc:
        print(f"    [ERR]  Failed to download {url}: {exc}")
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        return False


def process_copmgrr_list(copmgrr_list: list[int], output_dir: str) -> tuple[int, int]:
    """
    For each COPMGRR in *copmgrr_list*, check for new documents and download them
    into *output_dir*/COPMGRR{n}/.  Returns (files_downloaded, copmgrrs_updated).
    """
    total_files = total_copmgrrs = 0

    for copmgrr in copmgrr_list:
        print(f"\nCOPMGRR{copmgrr}  ->  {ISSUE_URL.format(n=copmgrr)}")
        links = get_document_links(copmgrr)
        time.sleep(REQUEST_DELAY)

        if not links:
            print("  (no downloadable documents found or page does not exist)")
            continue

        copmgrr_dir = os.path.join(output_dir, f"COPMGRR{copmgrr}")

        new_items = []
        for item in links:
            fname = sanitize(unquote(os.path.basename(item["url"].split("?")[0])))
            if not fname:
                fname = sanitize(item["label"]) + ".bin"
            fname = safe_fname(fname, copmgrr_dir)
            dest  = os.path.join(copmgrr_dir, fname)
            if not os.path.exists(dest):
                new_items.append((item, fname, dest))

        already = len(links) - len(new_items)
        if already:
            print(f"  {already} file(s) already up to date.")

        if not new_items:
            print("  Nothing new to download.")
            continue

        print(f"  {len(new_items)} new file(s) to download "
              f"(out of {len(links)} total).")
        os.makedirs(copmgrr_dir, exist_ok=True)
        total_copmgrrs += 1

        for item, fname, dest in new_items:
            if download_file(item["url"], dest):
                total_files += 1
            time.sleep(REQUEST_DELAY)

    return total_files, total_copmgrrs


def load_excel_copmgrrs(sheet_name: str) -> set[int]:
    """
    Read COPMGRR numbers from *sheet_name* in EXCEL_TRACKER.
    Returns an empty set if the file or sheet does not exist.
    """
    if not os.path.exists(EXCEL_TRACKER):
        return set()
    try:
        wb = openpyxl.load_workbook(EXCEL_TRACKER, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return set()
        ws = wb[sheet_name]
        numbers = set()
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header row
            if row[0] is not None:
                try:
                    numbers.add(int(row[0]))
                except (ValueError, TypeError):
                    pass
        wb.close()
        return numbers
    except Exception as exc:
        print(f"  [WARN] Could not read {EXCEL_TRACKER}: {exc}")
        return set()


def update_excel_copmgrrs(sheet_name: str, copmgrr_list: list[int]) -> None:
    """
    Overwrite *sheet_name* in EXCEL_TRACKER with the full *copmgrr_list*,
    sorted ascending.  Creates the workbook/sheet if they do not exist.
    """
    os.makedirs(os.path.dirname(EXCEL_TRACKER), exist_ok=True)
    if os.path.exists(EXCEL_TRACKER):
        wb = openpyxl.load_workbook(EXCEL_TRACKER)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append(["COPMGRR_Number", f"Last Updated: {timestamp}"])
    for n in sorted(copmgrr_list):
        ws.append([n])

    wb.save(EXCEL_TRACKER)
    print(f"  [Excel] '{sheet_name}' updated — {len(copmgrr_list)} COPMGRR(s) saved to {EXCEL_TRACKER} (as of {timestamp})")


def main():
    # ── Pending ───────────────────────────────────────────────────────────────
    if COPMGRR_NUMBERS:
        pending_list = COPMGRR_NUMBERS
        print("ERCOT COPMGRR Document Downloader  (manual list)")
    else:
        pending_list = get_pending_copmgrrs()
        if not pending_list:
            print("No pending COPMGRRs found.")
        else:
            print(f"Found {len(pending_list)} pending COPMGRR(s): "
                  f"{pending_list[0]} - {pending_list[-1]}")

    # ── Withdrawn ─────────────────────────────────────────────────────────────
    withdrawn_list = COPMGRR_NUMBERS if COPMGRR_NUMBERS else get_withdrawn_copmgrrs()
    withdrawn_new = withdrawn_list
    if withdrawn_list:
        print(f"Found {len(withdrawn_list)} withdrawn COPMGRR(s): "
              f"{withdrawn_list[0]} - {withdrawn_list[-1]}")
        tracked_withdrawn = load_excel_copmgrrs("List_Withdrawn")
        if tracked_withdrawn:
            withdrawn_new = [n for n in withdrawn_list if n not in tracked_withdrawn]
            print(f"  {len(tracked_withdrawn)} already tracked in Excel, "
                  f"{len(withdrawn_new)} new COPMGRR(s) to process.")
        else:
            print("  No prior tracking found — will process all withdrawn COPMGRRs.")

    # ── Approved ──────────────────────────────────────────────────────────────
    approved_list = COPMGRR_NUMBERS if COPMGRR_NUMBERS else get_approved_copmgrrs()
    approved_list = [n for n in approved_list if n > APPROVED_MIN_COPMGRR]
    approved_new = approved_list
    if approved_list:
        print(f"Found {len(approved_list)} approved COPMGRR(s) > {APPROVED_MIN_COPMGRR}: "
              f"{approved_list[0]} - {approved_list[-1]}")
        tracked_approved = load_excel_copmgrrs("List_Approved")
        if tracked_approved:
            approved_new = [n for n in approved_list if n not in tracked_approved]
            print(f"  {len(tracked_approved)} already tracked in Excel, "
                  f"{len(approved_new)} new COPMGRR(s) to process.")
        else:
            print("  No prior tracking found — will process all approved COPMGRRs.")

    print("=" * 60)

    # ── Process pending ───────────────────────────────────────────────────────
    p_files = p_copmgrrs = 0
    if pending_list:
        print(f"\n{'='*60}")
        print(f"PENDING COPMGRRs  ->  {os.path.abspath(PENDING_DIR)}")
        print(f"{'='*60}")
        p_files, p_copmgrrs = process_copmgrr_list(pending_list, PENDING_DIR)
        update_excel_copmgrrs("List_Pending", pending_list)

    # ── Process withdrawn ─────────────────────────────────────────────────────
    w_files = w_copmgrrs = 0
    if withdrawn_list:
        print(f"\n{'='*60}")
        print(f"WITHDRAWN COPMGRRs  ->  {os.path.abspath(WITHDRAWN_DIR)}")
        print(f"{'='*60}")
        w_files, w_copmgrrs = process_copmgrr_list(withdrawn_new, WITHDRAWN_DIR)
        update_excel_copmgrrs("List_Withdrawn", withdrawn_list)

    # ── Process approved ──────────────────────────────────────────────────────
    a_files = a_copmgrrs = 0
    if approved_list:
        print(f"\n{'='*60}")
        print(f"APPROVED COPMGRRs  ->  {os.path.abspath(APPROVED_DIR)}")
        print(f"{'='*60}")
        a_files, a_copmgrrs = process_copmgrr_list(approved_new, APPROVED_DIR)
        update_excel_copmgrrs("List_Approved", approved_list)

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print(f"Pending  : {p_files} new file(s) across {p_copmgrrs} COPMGRR(s)  ->  {os.path.abspath(PENDING_DIR)}")
    print(f"Withdrawn: {w_files} new file(s) across {w_copmgrrs} COPMGRR(s)  ->  {os.path.abspath(WITHDRAWN_DIR)}")
    print(f"Approved : {a_files} new file(s) across {a_copmgrrs} COPMGRR(s)  ->  {os.path.abspath(APPROVED_DIR)}")
    print(f"Total    : {p_files + w_files + a_files} new file(s) downloaded.")


if __name__ == "__main__":
    main()

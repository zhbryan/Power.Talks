"""
ERCOT NOGRR Document Downloader
================================
Automatically fetches all *pending*, *withdrawn*, *approved*, and *rejected* NOGRRs
from the ERCOT website, checks which documents are already saved locally,
and downloads only the new ones (PDF, DOC, DOCX, XLS, XLSX).

Usage
-----
    pip install requests beautifulsoup4 openpyxl
    python download_ercot_nogrr.py

Configuration
-------------
Edit the SETTINGS block below to change where files are saved, or to
override the auto-fetched lists with specific NOGRR numbers.
"""

import os
import re
import time
import requests
import openpyxl
from bs4 import BeautifulSoup
from datetime import datetime
from urllib.parse import urljoin, unquote

# ── SETTINGS ─────────────────────────────────────────────────────────────────

# Override: set to a list of NOGRR numbers to process instead of auto-fetching.
# Applies to all runs.  Leave as None for auto-fetch.
# Example: NOGRR_NUMBERS = [100, 110, 120]
NOGRR_NUMBERS = None

# Local folder where files will be saved (one sub-folder per NOGRR).
BASE_DIR = r"C:\Users\chunl\OneDrive\Documents\Business Ventures\Power.Talks\Documents Database\ERCOT.MKT.RULES\NOGRR"

# Seconds to pause between HTTP requests (be polite to the server).
REQUEST_DELAY = 1.5

# File extensions to download.
DOWNLOAD_EXTS = {".pdf", ".doc", ".docx", ".xls", ".xlsx"}

# Excel workbook used to track which NOGRR numbers have been seen before.
EXCEL_TRACKER = os.path.join(BASE_DIR, "Current List of NOGRRs.xlsx")

# ─────────────────────────────────────────────────────────────────────────────

BASE_URL           = "https://www.ercot.com"
NOGRR_LIST_URL     = "https://www.ercot.com/mktrules/issues/nogrr"
NOGRR_PENDING_URL  = "https://www.ercot.com/mktrules/issues/reports/nogrr/pending"
NOGRR_WITHDRAWN_URL = "https://www.ercot.com/mktrules/issues/reports/nogrr/withdrawn"
NOGRR_APPROVED_URL = "https://www.ercot.com/mktrules/issues/reports/nogrr/approved"
NOGRR_REJECTED_URL = "https://www.ercot.com/mktrules/issues/reports/nogrr/rejected"
ISSUE_URL          = "https://www.ercot.com/mktrules/issues/NOGRR{n}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; NOGRR-Downloader/1.0; "
        "non-commercial research)"
    )
}

MAX_PATH = 240  # conservative Windows path-length limit


def sanitize(name: str) -> str:
    """Strip characters that are illegal in file/folder names."""
    return re.sub(r'[\\/*?:"<>|]', "_", name).strip()


def safe_fname(fname: str, dest_dir: str) -> str:
    """Truncate fname so the full path stays within MAX_PATH."""
    stem, ext = os.path.splitext(fname)
    max_stem = MAX_PATH - len(dest_dir) - len(os.sep) - len(ext)
    if max_stem < 1:
        max_stem = 1
    return stem[:max_stem] + ext


def fetch_nogrr_numbers_from_report(url: str, label: str) -> list[int]:
    """Scrape an NOGRR report page and return sorted NOGRR numbers."""
    print(f"Fetching {label} NOGRR list from {url} ...")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [ERR] Could not fetch {label} NOGRR list: {exc}")
        return []

    numbers = re.findall(r"/mktrules/issues/NOGRR(\d+)", resp.text, re.IGNORECASE)
    return sorted(set(int(n) for n in numbers))


def get_pending_nogrrs() -> list[int]:
    """
    Scrape the ERCOT NOGRR listing page for the Pending section,
    falling back to the pending report page.
    """
    print(f"Fetching pending NOGRR list from {NOGRR_LIST_URL} ...")
    try:
        resp = requests.get(NOGRR_LIST_URL, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [ERR] Could not fetch NOGRR list: {exc}")
        return fetch_nogrr_numbers_from_report(NOGRR_PENDING_URL, "pending")

    soup = BeautifulSoup(resp.text, "html.parser")

    pending_heading = soup.find(
        lambda tag: tag.name in ("h3", "h4", "h5")
        and "pending" in tag.get_text(strip=True).lower()
    )

    nogrr_numbers = []
    if pending_heading:
        for sibling in pending_heading.find_next_siblings():
            if sibling.name in ("h3", "h4", "h5"):
                break
            for a in sibling.find_all("a", href=True):
                m = re.search(r"/mktrules/issues/NOGRR(\d+)", a["href"], re.IGNORECASE)
                if m:
                    nogrr_numbers.append(int(m.group(1)))
    else:
        matches = re.findall(r"/mktrules/issues/NOGRR(\d+)", resp.text, re.IGNORECASE)
        nogrr_numbers = [int(n) for n in matches]

    if not nogrr_numbers:
        print("  [INFO] No pending NOGRRs found on main page; trying report page.")
        return fetch_nogrr_numbers_from_report(NOGRR_PENDING_URL, "pending")

    seen, unique = set(), []
    for n in nogrr_numbers:
        if n not in seen:
            seen.add(n)
            unique.append(n)
    return unique


def get_document_links(nogrr: int) -> list[dict]:
    """
    Fetch the ERCOT issue page for one NOGRR and return a list of dicts
    {'label': str, 'url': str} for every downloadable file found.
    Returns an empty list if the page is unreachable or has no files.
    """
    url = ISSUE_URL.format(n=nogrr)
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        if resp.status_code == 404:
            return []
        resp.raise_for_status()
    except requests.RequestException as exc:
        print(f"  [WARN] Could not fetch {url}: {exc}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    links, seen = [], set()

    for a in soup.find_all("a", href=True):
        href = a["href"]
        ext  = os.path.splitext(href.split("?")[0])[-1].lower()
        if ext not in DOWNLOAD_EXTS:
            continue
        full_url = urljoin(BASE_URL, href)
        if full_url in seen:
            continue
        seen.add(full_url)
        label = a.get_text(separator=" ", strip=True) or os.path.basename(href)
        links.append({"label": label, "url": full_url})

    return links


def download_file(url: str, dest_path: str) -> bool:
    """Download *url* to *dest_path*. Returns True on success.

    Writes to a temporary file first and renames on success so that a failed
    or interrupted download never leaves a corrupt file that would be mistaken
    for a completed download on the next run.
    """
    tmp_path = dest_path + ".tmp"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=60, stream=True)
        resp.raise_for_status()
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        with open(tmp_path, "wb") as fh:
            for chunk in resp.iter_content(chunk_size=65536):
                fh.write(chunk)
        os.replace(tmp_path, dest_path)
        size_kb = os.path.getsize(dest_path) / 1024
        print(f"    [OK]   {os.path.basename(dest_path)}  ({size_kb:.1f} KB)")
        return True
    except (requests.RequestException, OSError) as exc:
        print(f"    [ERR]  Failed to download {url}: {exc}")
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        return False


def process_nogrr_list(nogrr_list: list[int], output_dir: str) -> tuple[int, int]:
    """
    For each NOGRR in *nogrr_list*, check for new documents and download them
    into *output_dir*/NOGRR{n}/.  Returns (files_downloaded, nogrrs_updated).
    """
    total_files = total_nogrrs = 0

    for nogrr in nogrr_list:
        print(f"\nNOGRR{nogrr}  ->  {ISSUE_URL.format(n=nogrr)}")
        links = get_document_links(nogrr)
        time.sleep(REQUEST_DELAY)

        if not links:
            print("  (no downloadable documents found or page does not exist)")
            continue

        nogrr_dir = os.path.join(output_dir, f"NOGRR{nogrr}")

        new_items = []
        for item in links:
            fname = sanitize(unquote(os.path.basename(item["url"].split("?")[0])))
            if not fname:
                fname = sanitize(item["label"]) + ".bin"
            fname = safe_fname(fname, nogrr_dir)
            dest  = os.path.join(nogrr_dir, fname)
            if not os.path.exists(dest):
                new_items.append((item, fname, dest))

        already = len(links) - len(new_items)
        if already:
            print(f"  {already} file(s) already up to date.")

        if not new_items:
            print("  Nothing new to download.")
            continue

        print(f"  {len(new_items)} new file(s) to download "
              f"(out of {len(links)} total).")
        os.makedirs(nogrr_dir, exist_ok=True)
        total_nogrrs += 1

        for item, fname, dest in new_items:
            if download_file(item["url"], dest):
                total_files += 1
            time.sleep(REQUEST_DELAY)

    return total_files, total_nogrrs


def load_excel_nogrrs(sheet_name: str) -> set[int]:
    """
    Read NOGRR numbers from *sheet_name* in EXCEL_TRACKER.
    Returns an empty set if the file or sheet does not exist.
    """
    if not os.path.exists(EXCEL_TRACKER):
        return set()
    try:
        wb = openpyxl.load_workbook(EXCEL_TRACKER, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return set()
        ws = wb[sheet_name]
        numbers = set()
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header row
            if row[0] is not None:
                try:
                    numbers.add(int(row[0]))
                except (ValueError, TypeError):
                    pass
        wb.close()
        return numbers
    except Exception as exc:
        print(f"  [WARN] Could not read {EXCEL_TRACKER}: {exc}")
        return set()


def update_excel_nogrrs(sheet_name: str, nogrr_list: list[int]) -> None:
    """
    Overwrite *sheet_name* in EXCEL_TRACKER with the full *nogrr_list*,
    sorted ascending.  Creates the workbook/sheet if they do not exist.
    """
    os.makedirs(os.path.dirname(EXCEL_TRACKER), exist_ok=True)
    if os.path.exists(EXCEL_TRACKER):
        wb = openpyxl.load_workbook(EXCEL_TRACKER)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append(["NOGRR_Number", f"Last Updated: {timestamp}"])
    for n in sorted(nogrr_list):
        ws.append([n])

    wb.save(EXCEL_TRACKER)
    print(f"  [Excel] '{sheet_name}' updated — {len(nogrr_list)} NOGRR(s) saved to {EXCEL_TRACKER} (as of {timestamp})")


def resolve_new(full_list: list[int], sheet_name: str, label: str) -> list[int]:
    """Compare *full_list* against Excel tracker and return only new NOGRRs."""
    tracked = load_excel_nogrrs(sheet_name)
    if tracked:
        new = [n for n in full_list if n not in tracked]
        print(f"  {len(tracked)} already tracked in Excel, "
              f"{len(new)} new NOGRR(s) to process.")
    else:
        new = full_list
        print(f"  No prior tracking found — will process all {label} NOGRRs.")
    return new


def main():
    # ── Pending ───────────────────────────────────────────────────────────────
    if NOGRR_NUMBERS:
        pending_list = NOGRR_NUMBERS
        print("ERCOT NOGRR Document Downloader  (manual list)")
    else:
        pending_list = get_pending_nogrrs()
        if not pending_list:
            print("No pending NOGRRs found.")
        else:
            print(f"Found {len(pending_list)} pending NOGRR(s): "
                  f"{pending_list[0]} - {pending_list[-1]}")

    # ── Withdrawn ─────────────────────────────────────────────────────────────
    withdrawn_list = NOGRR_NUMBERS if NOGRR_NUMBERS else fetch_nogrr_numbers_from_report(NOGRR_WITHDRAWN_URL, "withdrawn")
    withdrawn_new = []
    if withdrawn_list:
        print(f"Found {len(withdrawn_list)} withdrawn NOGRR(s): "
              f"{withdrawn_list[0]} - {withdrawn_list[-1]}")
        withdrawn_new = resolve_new(withdrawn_list, "List_Withdrawn", "withdrawn")

    # ── Approved ──────────────────────────────────────────────────────────────
    approved_list = NOGRR_NUMBERS if NOGRR_NUMBERS else fetch_nogrr_numbers_from_report(NOGRR_APPROVED_URL, "approved")
    approved_new = []
    if approved_list:
        print(f"Found {len(approved_list)} approved NOGRR(s): "
              f"{approved_list[0]} - {approved_list[-1]}")
        approved_new = resolve_new(approved_list, "List_Approved", "approved")

    # ── Rejected ──────────────────────────────────────────────────────────────
    rejected_list = NOGRR_NUMBERS if NOGRR_NUMBERS else fetch_nogrr_numbers_from_report(NOGRR_REJECTED_URL, "rejected")
    rejected_new = []
    if rejected_list:
        print(f"Found {len(rejected_list)} rejected NOGRR(s): "
              f"{rejected_list[0]} - {rejected_list[-1]}")
        rejected_new = resolve_new(rejected_list, "List_Rejected", "rejected")

    print("=" * 60)

    # ── Process pending ───────────────────────────────────────────────────────
    p_files = p_nogrrs = 0
    if pending_list:
        print(f"\n{'='*60}")
        print(f"PENDING NOGRRs  ->  {os.path.abspath(BASE_DIR)}")
        print(f"{'='*60}")
        p_files, p_nogrrs = process_nogrr_list(pending_list, BASE_DIR)
        update_excel_nogrrs("List_Pending", pending_list)

    # ── Process withdrawn ─────────────────────────────────────────────────────
    w_files = w_nogrrs = 0
    if withdrawn_list:
        print(f"\n{'='*60}")
        print(f"WITHDRAWN NOGRRs  ->  {os.path.abspath(BASE_DIR)}")
        print(f"{'='*60}")
        w_files, w_nogrrs = process_nogrr_list(withdrawn_new, BASE_DIR)
        update_excel_nogrrs("List_Withdrawn", withdrawn_list)

    # ── Process approved ──────────────────────────────────────────────────────
    a_files = a_nogrrs = 0
    if approved_list:
        print(f"\n{'='*60}")
        print(f"APPROVED NOGRRs  ->  {os.path.abspath(BASE_DIR)}")
        print(f"{'='*60}")
        a_files, a_nogrrs = process_nogrr_list(approved_new, BASE_DIR)
        update_excel_nogrrs("List_Approved", approved_list)

    # ── Process rejected ──────────────────────────────────────────────────────
    r_files = r_nogrrs = 0
    if rejected_list:
        print(f"\n{'='*60}")
        print(f"REJECTED NOGRRs  ->  {os.path.abspath(BASE_DIR)}")
        print(f"{'='*60}")
        r_files, r_nogrrs = process_nogrr_list(rejected_new, BASE_DIR)
        update_excel_nogrrs("List_Rejected", rejected_list)

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print(f"Pending  : {p_files} new file(s) across {p_nogrrs} NOGRR(s)  ->  {os.path.abspath(BASE_DIR)}")
    print(f"Withdrawn: {w_files} new file(s) across {w_nogrrs} NOGRR(s)  ->  {os.path.abspath(BASE_DIR)}")
    print(f"Approved : {a_files} new file(s) across {a_nogrrs} NOGRR(s)  ->  {os.path.abspath(BASE_DIR)}")
    print(f"Rejected : {r_files} new file(s) across {r_nogrrs} NOGRR(s)  ->  {os.path.abspath(BASE_DIR)}")
    print(f"Total    : {p_files + w_files + a_files + r_files} new file(s) downloaded.")


if __name__ == "__main__":
    main()

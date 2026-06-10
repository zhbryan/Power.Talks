#!/usr/bin/env python3
"""
ERCOT NOGRR Profile Creator
Implements the ERCOT Market Rules Profile skill for all NOGRR issues.
Reads each -01 document, extracts 18 profile fields, saves JSON to
<ISSUE_ID>/Quick runs/<ISSUE_ID> Profile.json
"""

import os, re, json
from datetime import datetime
from docx import Document
import openpyxl
import win32com.client

CATEGORY = "NOGRR"

# ─── TIMELINE PATTERNS ───────────────────────────────────────────────────────
_TL_BALLOT_PAT = re.compile(r'(ros|tac|prs|rms)\s*(?:email\s*)?ballot', re.IGNORECASE)
_TL_REPORT_PAT = re.compile(r'(ros|tac|prs|rms)\s*report', re.IGNORECASE)
_TL_BOARD_PAT  = re.compile(r'board.?report', re.IGNORECASE)
_TL_PUCT_PAT   = re.compile(r'puct.?(?:report|filing)', re.IGNORECASE)
_TL_ERCOT_PAT  = re.compile(r'ercot.?comments', re.IGNORECASE)
_TL_IMPACT_PAT = re.compile(r'impact.?analysis', re.IGNORECASE)

# ─── SETTINGS ────────────────────────────────────────────────────────────────
BASE_DIR      = r"E:\wamp64\www\Power.Talks\Documents Database\ERCOT.MKT.RULES\NOGRR"
EXCEL_TRACKER = os.path.join(BASE_DIR, "Current List of NOGRRs.xlsx")

# ─── STATUS MAP ──────────────────────────────────────────────────────────────
def load_status_map():
    status_map = {}
    try:
        wb = openpyxl.load_workbook(EXCEL_TRACKER, read_only=True, data_only=True)
        for sheet, status in [("List_Pending", "Pending"),
                               ("List_Withdrawn", "Withdrawn"),
                               ("List_Approved", "Approved"),
                               ("List_Rejected", "Rejected")]:
            if sheet in wb.sheetnames:
                for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:
                        try:
                            status_map[int(row[0])] = status
                        except (ValueError, TypeError):
                            pass
    except Exception as e:
        print(f"  Warning: Excel tracker read failed: {e}")
    return status_map

# ─── DOCUMENT READERS ────────────────────────────────────────────────────────
def tables_from_docx(path):
    doc = Document(path)
    return [[[c.text.strip() for c in row.cells] for row in tbl.rows]
            for tbl in doc.tables]

def tables_from_doc(path):
    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False
    try:
        d = word.Documents.Open(os.path.abspath(path), ReadOnly=True)
        result = []
        for ti in range(1, d.Tables.Count + 1):
            tbl = d.Tables(ti)
            rows = []
            for r in range(1, tbl.Rows.Count + 1):
                cells = []
                for c in range(1, tbl.Columns.Count + 1):
                    try:
                        cells.append(tbl.Cell(r, c).Range.Text
                                     .replace('\r', ' ').replace('\x07', '').strip())
                    except Exception:
                        cells.append('')
                rows.append(cells)
            result.append(rows)
        d.Close(False)
        return result
    finally:
        word.Quit()

# ─── DATE NORMALIZER ─────────────────────────────────────────────────────────
_MONTH_FMTS = ['%B %d, %Y', '%B %d %Y', '%b %d, %Y', '%b %d %Y',
               '%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d']

def date_from_fname(fname):
    stem = os.path.splitext(fname)[0]
    for pat in [r'[_\-](\d{6}|\d{8})(?:\s|$)', r'\s(\d{6}|\d{8})\s*$', r'(\d{8})']:
        m = re.search(pat, stem)
        if m:
            d = m.group(1)
            try:
                if len(d) == 6:
                    return datetime.strptime(d, '%m%d%y').strftime('%Y-%m-%d')
                else:
                    return datetime.strptime(d, '%m%d%Y').strftime('%Y-%m-%d')
            except ValueError:
                pass
    return None

def parse_reason_list(text):
    if not text:
        return []
    parts = re.split(r'[\x01\x02\x03]+', text)
    result = []
    for p in parts:
        p = p.strip().rstrip('.')
        p = re.sub(r'^Other\s*:\s*\(explain\)\s*', 'Other: ', p, flags=re.IGNORECASE)
        p = re.sub(r'^\s*[\-\*•]\s*', '', p).strip()
        if p and len(p) > 3:
            result.append(p)
    return result

def build_profile_timeline(folder, date_posted):
    events = []
    if date_posted:
        events.append({'date': date_posted, 'event': f'{CATEGORY} Posted', 'doc': 'Initial posting'})
    for fname in sorted(os.listdir(folder)):
        fpath = os.path.join(folder, fname)
        if os.path.isdir(fpath):
            continue
        fl = fname.lower()
        date_str = date_from_fname(fname) or 'Unknown'
        stem = os.path.splitext(fname)[0]
        if _TL_BALLOT_PAT.search(fl):
            m = _TL_BALLOT_PAT.search(fl)
            body = m.group(1).upper()
            events.append({'date': date_str, 'event': f'{body} Ballot', 'doc': stem})
        elif _TL_BOARD_PAT.search(fl):
            events.append({'date': date_str, 'event': 'Board Report', 'doc': stem})
        elif _TL_PUCT_PAT.search(fl):
            events.append({'date': date_str, 'event': 'PUCT Filing', 'doc': stem})
        elif _TL_REPORT_PAT.search(fl):
            m = _TL_REPORT_PAT.search(fl)
            body = m.group(1).upper()
            events.append({'date': date_str, 'event': f'{body} Report', 'doc': stem})
        elif _TL_ERCOT_PAT.search(fl):
            events.append({'date': date_str, 'event': 'ERCOT Comments', 'doc': stem})
        elif _TL_IMPACT_PAT.search(fl):
            events.append({'date': date_str, 'event': 'Impact Analysis', 'doc': stem})
    events.sort(key=lambda x: (x['date'] if x['date'] != 'Unknown' else '9999-99-99', x['event']))
    return events

def normalize_date(text):
    if not text:
        return None
    text = text.strip()
    for fmt in _MONTH_FMTS:
        try:
            return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    m = re.search(r'(\w+\s+\d{1,2},?\s+\d{4})', text)
    if m:
        s = re.sub(r',', '', m.group(1))
        for fmt in ['%B %d %Y', '%b %d %Y']:
            try:
                return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
            except ValueError:
                pass
    return text

# ─── FIELD EXTRACTOR ─────────────────────────────────────────────────────────
def first_nonempty_sibling(row, label_idx):
    label = row[label_idx].lower()
    for j in range(label_idx + 1, len(row)):
        v = row[j].strip()
        if v and v.lower() != label:
            return v
    return None

def extract_fields_from_tables(tables):
    fields = {
        'title': None, 'date_posted_decision': None,
        'timeline_requested_resolution': None, 'effective_date': None,
        'governing_document_sections': [], 'related_documents_requiring_revision': [],
        'revision_description': None, 'reason_for_revision': [],
        'business_case': None, 'sponsor_name': None, 'sponsor_email': None,
        'sponsor_company': None, 'sponsor_phone': None, 'market_segment': None,
    }
    if not tables:
        return fields

    for row in tables[0]:
        for i, cell in enumerate(row):
            cl = cell.strip().lower()
            if not cl:
                continue
            v = first_nonempty_sibling(row, i)
            if not v:
                continue
            if 'nogrr title' in cl or cl == 'title':
                fields['title'] = fields['title'] or v
            elif cl.startswith('date posted') or 'decision' in cl:
                fields['date_posted_decision'] = fields['date_posted_decision'] or normalize_date(v)
            elif 'requested resolution' in cl:
                fields['timeline_requested_resolution'] = fields['timeline_requested_resolution'] or v
            elif 'operating guide sections' in cl or 'sections requiring' in cl or 'nodal operating' in cl:
                if not fields['governing_document_sections']:
                    fields['governing_document_sections'] = [
                        s.strip() for s in re.split(r'[;,\n]', v) if s.strip()
                    ]
            elif 'related documents' in cl or 'related revision' in cl:
                if not fields['related_documents_requiring_revision']:
                    items = [s.strip() for s in re.split(r'[;,\n]', v)
                             if s.strip() and s.strip().lower() not in ('none', 'n/a')]
                    fields['related_documents_requiring_revision'] = items
            elif 'revision description' in cl:
                fields['revision_description'] = fields['revision_description'] or v
            elif 'reason for revision' in cl:
                if not fields['reason_for_revision']:
                    fields['reason_for_revision'] = parse_reason_list(v)
            elif 'business case' in cl or 'justification' in cl or 'market impacts' in cl:
                fields['business_case'] = fields['business_case'] or v

    if len(tables) > 1:
        for row in tables[1]:
            for i, cell in enumerate(row):
                cl = cell.strip().lower()
                if not cl:
                    continue
                v = first_nonempty_sibling(row, i)
                if not v:
                    continue
                if cl == 'name':
                    fields['sponsor_name'] = fields['sponsor_name'] or v
                elif 'e-mail' in cl or 'email' in cl:
                    fields['sponsor_email'] = fields['sponsor_email'] or v
                elif cl == 'company':
                    fields['sponsor_company'] = fields['sponsor_company'] or v
                elif 'phone' in cl or 'cell' in cl:
                    if not fields['sponsor_phone']:
                        fields['sponsor_phone'] = v
                elif 'market segment' in cl:
                    fields['market_segment'] = fields['market_segment'] or v
    return fields

# ─── PROFILE BUILDER ─────────────────────────────────────────────────────────
def find_main_doc(folder):
    for fname in sorted(os.listdir(folder)):
        if re.search(r'-0*1[\b\s\-_]', fname) or re.search(r'-01\.', fname):
            ext = os.path.splitext(fname)[1].lower()
            if ext in ('.docx', '.doc'):
                return os.path.join(folder, fname), ext
    return None, None

def build_profile(folder, issue_num, status):
    issue_id = f"NOGRR{issue_num}"
    profile = {
        "category": CATEGORY,
        "issue_id": issue_id,
        "issue_number": issue_num,
        "title": None,
        "date_posted_decision": None,
        "timeline_requested_resolution": None,
        "status": status,
        "effective_date": None,
        "governing_document_sections": [],
        "related_documents_requiring_revision": [],
        "revision_description": None,
        "reason_for_revision": [],
        "business_case": None,
        "sponsor_name": None,
        "sponsor_email": None,
        "sponsor_company": None,
        "sponsor_phone": None,
        "market_segment": None,
        "timeline": [],
    }

    path, ext = find_main_doc(folder)
    if path:
        try:
            tables = tables_from_docx(path) if ext == '.docx' else tables_from_doc(path)
            fields = extract_fields_from_tables(tables)
            profile.update({k: v for k, v in fields.items() if v})
        except Exception as e:
            print(f"  Warning parsing {os.path.basename(path)}: {e}")

    profile["timeline"] = build_profile_timeline(folder, profile.get("date_posted_decision"))
    return profile

# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    status_map = load_status_map()

    folders = sorted([
        d for d in os.listdir(BASE_DIR)
        if os.path.isdir(os.path.join(BASE_DIR, d)) and re.match(r'NOGRR\d+', d)
    ], key=lambda x: int(re.search(r'\d+', x).group()))

    print(f"Processing {len(folders)} NOGRR folders...\n")
    ok = err = 0

    for folder_name in folders:
        n = int(re.search(r'\d+', folder_name).group())
        folder = os.path.join(BASE_DIR, folder_name)
        status = status_map.get(n, "Unknown")

        try:
            profile = build_profile(folder, n, status)
            quick = os.path.join(folder, "Quick runs")
            os.makedirs(quick, exist_ok=True)
            out = os.path.join(quick, f"NOGRR{n} Profile.json")
            with open(out, 'w', encoding='utf-8') as f:
                json.dump(profile, f, indent=2, ensure_ascii=False)
            title = profile['title'] or '(title not found)'
            print(f"[OK] NOGRR{n:>4}  {status:<10}  {title[:65]}")
            ok += 1
        except Exception as e:
            print(f"[ERR] NOGRR{n}: {e}")
            err += 1

    print(f"\nDone: {ok} profiles created, {err} errors.")

if __name__ == "__main__":
    main()
